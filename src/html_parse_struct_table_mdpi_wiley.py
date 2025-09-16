# -*- coding: utf-8 -*-
"""
Created on Tue Nov 28 15:34:27 2023

@author: dell
"""

from bs4 import BeautifulSoup
import os
import json
from openpyxl import load_workbook,Workbook
from prettytable import PrettyTable
from doc_parser import parser

def print_table(table):
    x=PrettyTable()
    if len(table['head'])==0:
        ncol=len(table['body'][0])
        fn=[]
        for j in range(0,ncol):
            fn.append(str(j))
        x.field_names=fn  
    else:
        if isinstance(table['head'][0],str):
            try:
                x.field_names=table['head']
            except:
                ncol=len(table['head'])
                fn=[]
                for j in range(0,ncol):
                    fn.append(str(j))
                x.field_names=fn
                x.add_row(table['head'])
        elif isinstance(table['head'][0],list):
            try:
                x.field_names=table['head'][0]
            except:
                ncol=len(table['head'][0])
                fn=[]
                for j in range(0,ncol):
                    fn.append(str(j))
                x.field_names=fn  
                x.add_row(table['head'][0])
            for j in range(1,len(table['head'])):
                x.add_row(table['head'][j])
    for body in table['body']:
        x.add_row(body)
    print(x)    

def join_table(tables):
    tables1=[]
    ifjoin=[0,]
    rowmark=[0,] # 1: merge row mark; 0. not merge row mark
    for i in range(1,len(tables)):
        tab0=tables[i-1]
        tab1=tables[i]
        if (len(tab0['head'])==len(tab1['head']))and(len(tab0['body'])==len(tab1['body'])):
            ifjoin.append(1)
            # examine if the first column of two tables is the same. if the same, then merge
            flag=1
            for j in range(0,len(tab0['head'])):
                if not tab0['head'][j][0]==tab1['head'][j][0]:
                    flag=0
                    break
            for j in range(0,len(tab0['body'])):
                if not tab0['body'][j][0]==tab1['body'][j][0]:
                    flag=0
                    break  
            # examine if the first column of the second table is the empty. if the empty, then merge
            flag2=1
            for j in range(0,len(tab1['head'])):
                if not tab1['head'][j][0]=='':
                    flag2=0
                    break
            for j in range(0,len(tab1['body'])):
                if not tab1['body'][j][0]=='':
                    flag2=0
                    break             
            rowmark.append( (flag or flag2) )
        else:
            ifjoin.append(0)
            rowmark.append(0)
            
    for i in range(0,len(ifjoin)):    # loop ober tables
        if not ifjoin[i]:
            tables1.append(tables[i])
        else:
            tab=tables[i]['head']
            for j in range(0,len(tab)):   # loop over rows
                for k in range(rowmark[i],len(tab[j])):    # loop over column
                    tables1[-1]['head'][j].append(tab[j][k])
            tab=tables[i]['body']
            for j in range(0,len(tab)):   # loop over rows
                for k in range(rowmark[i],len(tab[j])):    # loop over column
                    tables1[-1]['body'][j].append(tab[j][k])            
    return tables1


def fill_column(table,maxcol):
    for i in range(0,len(table)):
        while len(table[i])<maxcol:
            table[i].append('')
    return table

def get_max_column(table):
    maxcol=0
    for i in range(0,len(table)):
        l=len(table[i])
        if l>maxcol:
            maxcol=l
    return maxcol

def add_data(data,table,nr,nc,rs,cs):
    row=table[nr]
    n=len(row)
    if n>nc+1:
        for i in range(nc,n):
            if row[i]=='':
               nc=i
               break
    # enrich the table
    for i in range(nr,nr+rs):
        while len(table)<i+1:
            table.append([])
        for j in range(nc,nc+cs):
            while len(table[i])<j+1:
                table[i].append('')
            table[i][j]=data
    
    maxcol=get_max_column(table)
    
    table=fill_column(table,maxcol)        
    
    return table,nc+cs
            
def message(msg):            
    global log,iflog
    print(msg)
    if iflog:
        if len(msg)==0:
            return
        if not msg[-1]=='\n':
            log.write(msg+'\n')
        else:
            log.write(msg)
    
def parse_rows(inp,mtype):
    
    data=[]
    nr=0   
    rows=inp.find_all('tr')
    entity_attr=['class','rowspan','colspan','style','align','valign']
    for r in rows:
        nc=0
        for key in r.attrs.keys():
            if not key in entity_attr:
                message(key)
        if len(data)<nr+1:
            data.append([])
        cols=r.find_all(mtype)       
        for c in cols: 
            for key in c.attrs.keys():  
                if not key in entity_attr:
                    message(key)            
            text=c.text
            text=text.replace('\n',' ').strip()
            rs=1
            cs=1
            if c.has_attr('rowspan'):
                rs=int(c.attrs['rowspan'])
            if c.has_attr('colspan'):
                cs=int(c.attrs['colspan'])  
            data,nc=add_data(text,data,nr,nc,rs,cs)         
            
        nr=nr+1
    return data
        

def parse_single_table(tabs,title):
    tables=[]
    ntab=len(tabs)
    if ntab>1:
        message('WARNING: multiple table in '+title)
    elif ntab==0:
        message('WARNING: zero table in '+title)
        return tables
    
    for i in range(0,len(tabs)):
        nohead=0
        head=[]
        body=[]
        tables.append({})
        tables[-1]['title']=title
        tables[-1]['head']=head
        tables[-1]['body']=body         
        thead=tabs[i].find_all('thead')
        nthead=len(thead)
        if nthead>1:
            message('WARNING: multiple table head in '+title)
        elif nthead==0:
            nohead=1
            #message('WARNING: zero table head in a springer html')
        if nthead>=1:
            head=parse_rows(thead[0],'th')
        
        tbody=tabs[i].find_all('tbody')
        ntbody=len(tbody)
        if ntbody>1:
            message('WARNING: multiple table body in '+title)
        elif ntbody==0:
            message('WARNING: zero table body in '+title)
        if ntbody>=1:    
            body=parse_rows(tbody[0],'td')
        
        # make head and body have the same column
        maxcol_head=get_max_column(head)
        maxcol_body=get_max_column(body)
        if maxcol_head>maxcol_body:
            message('WARNING: column_head > column_body')
            body=fill_column(body,maxcol_head)
        elif (not nohead) and (maxcol_body>maxcol_head):
            message('WARNING: column_body > column_head')
            head=fill_column(head,maxcol_body)
            
        tables[-1]['title']=title
        tables[-1]['head']=head
        tables[-1]['body']=body  
    
    if ntab>1:
        tables=join_table(tables)        
        
    return tables

def parse_table_springer(doc):
    
    # get title
    div=doc.find('div',id='main-content')
    tmp=div.find_all('h1')
    title=tmp[0].text
    tables=[] 
    
    # get table
    tabs=div.find_all('table')
    ntab=len(tabs)
    if ntab>1:
        message('WARNING: multiple table in a springer html')
    elif ntab==0:
        message('WARNING: zero table in a springer html')
        return tables
    
    for tab in tabs:
        nohead=0
        head=[]
        body=[]
        tables.append({})
        tables[-1]['title']=title
        tables[-1]['head']=head
        tables[-1]['body']=body         
        thead=tab.find_all('thead')
        nthead=len(thead)
        if nthead>1:
            message('WARNING: multiple table head in a table')
        elif nthead==0:
            nohead=1
            #message('WARNING: zero table head in a springer html')
        if nthead>=1:
            head=parse_rows(thead[0],'th')
        
        tbody=tab.find_all('tbody')
        ntbody=len(tbody)
        if ntbody>1:
            message('WARNING: multiple table body in a table')
        elif ntbody==0:
            message('WARNING: zero table body in a springer html')
        if ntbody>=1:    
            body=parse_rows(tbody[0],'td')
        
        # make head and body have the same column
        maxcol_head=get_max_column(head)
        maxcol_body=get_max_column(body)
        if maxcol_head>maxcol_body:
            message('WARNING: column_head > column_body')
            body=fill_column(body,maxcol_head)
        elif (not nohead) and (maxcol_body>maxcol_head):
            message('WARNING: column_body > column_head')
            head=fill_column(head,maxcol_body)
            
        tables[-1]['title']=title
        tables[-1]['head']=head
        tables[-1]['body']=body  
    
    if ntab>1:
        tables=join_table(tables)    
    
    return tables
    

def parse_table_wiley(doc):
    tables=[]
    # get title
    divs=doc.find_all('div',class_='article-table-content')
    for i in range(0,len(divs)):
        title=''
        tmp=divs[i].find_all('header')
        nheader=len(tmp)
        if nheader>1:
            message('WARNING: multiple table title in '+title)
        elif nheader==0:
            message('WARNING: zero table title in '+title)  
        for child in tmp[0].children:
            title=title+child.text.strip()
        
        # get table
        tabs=divs[i].find_all('table')
        tables.extend(parse_single_table(tabs,title))
    
    return tables

def parse_table_mdpi(doc):
    tables=[]
    # get title
    divs=doc.find_all('div',class_='html-table_show')
    for i in range(0,len(divs)):
        title=''
        tmp=divs[i].find_all('div',class_='html-caption')
        nheader=len(tmp)
        if nheader==0:
            tmp=divs[i].find_all('caption')
            nheader=len(tmp)
        if nheader>1:
            message('WARNING: multiple table title in '+title)
        elif nheader==0:
            message('WARNING: zero table title in '+title)  
        for child in tmp[0].children:
            title=title+child.text.strip()
        
        # get table
        tabs=divs[i].find_all('table')
        tables.extend(parse_single_table(tabs,title))
    
    return tables

def get_doi_dict():
    doi_dict={}
    wospath="D:/Tsinghua/Project/Fatigue Data Framework/search data/wos/20220917/"
    fileName="AM_fatigue_adddoi.xlsx"
    sheetName="savedrecs"
    if os.path.exists(wospath+'doi_dict.json'):
        f=open(wospath+'doi_dict.json','r')
        doi_dict=json.load(f)
        return doi_dict
    wb=load_workbook(filename=wospath+fileName)
    ws=wb[sheetName]
    nrow=ws.max_row
    ncol=ws.max_column
    for i in range(1,ncol+1):
        if ws.cell(1,i).value=='DOI':
            col=i
            break
    for i in range(2,nrow+1):
        doi=ws.cell(i,col).value.strip('\n')
        doi_trans=doi.replace('/','_').replace(':','_')
        doi_dict[doi_trans]=doi

    return doi_dict

def process_springer():
    global log,iflog
    iflog=1
    path='D:/glass/html_springer/'
    logpath=path+'parse.log'
    

    #doi_dict=get_doi_dict()
    filelist=os.listdir(path)
    doc_dict={}
    
    #dirlist=['10.1007_s11665-016-2140-2',]
    log=open(logpath,'w')
    
    for file in filelist:
        leng=len(file)
        tables=[]


            #

        if leng>5 and file[leng-5:leng]=='.html':
            filename=file[0:leng-5]
            doi=filename.replace('_','/')
            message(path+file)
            try:
                f=open(path+file,'r',encoding='utf-8')
                doc=BeautifulSoup(f,"lxml")
                f.close()
                tables.extend(parse_table_springer(doc))
                ifsuccess=1
            except:
                ifsuccess=0
            if not ifsuccess:
                message("PRASE ERROR")
            doc={}
            doc['path']=path
            doc['file']=file
            doc['doi']=doi
            doc['content']=parser(path,file,'springer',doi).sections
            doc['tables']=tables
            doc_dict[doi]=doc
    log.close()

    return doc_dict

def process_wiley():
    global log,iflog
    iflog=1
    #path='D:/glass/html_wiley/'
    path='D:/glass-total-results/html_wiley/'
    logpath=path+'parse.log'
    #doi_dict=get_doi_dict()
    filelist=os.listdir(path)
    doc_dict={}

    #dirlist=['10.1007_s11665-016-2140-2',]
    log=open(logpath,'w')
    #filelist=os.listdir(path)


    #doi=dir0.replace('_','/')
    for file in filelist:
        leng=len(file)
        tables=[]
        if leng>5 and file[leng-5:leng]=='.html':
            filename=file[0:leng-5]
            #try:
                #doi=doi_dict[filename]
            #except:
            doi=filename.replace('_','/')            
            message(path+file)
            try:
                f=open(path+file,'r',encoding='utf-8')
                doc=BeautifulSoup(f,"lxml")
                f.close()
                tables.extend(parse_table_wiley(doc))
                ifsuccess=1
            except:
                ifsuccess=0
            if not ifsuccess:
                message("PRASE ERROR")
            doc={}
            doc['path']=path
            doc['file']=file
            doc['doi']=doi
            doc['content']=parser(path,file,'wiley',doi).sections
            doc['tables']=tables
            doc_dict[doi]=doc
    log.close()
    return doc_dict


def process_mdpi():
    global log,iflog
    iflog=1
    #path='D:/glass/html_mdpi/'
    path='D:/glass-total-results/html_mdpi/'
    logpath=path+'parse.log'
    #doi_dict=get_doi_dict()
    filelist=os.listdir(path)
    doc_dict={}
    
    log=open(logpath,'w')
    filelist=os.listdir(path)
    
    for file in filelist:
        leng=len(file)
        tables=[]
        if leng>5 and file[leng-5:leng]=='.html':
            filename=file[0:leng-5]
                #doi=doi_dict[filename]
            doi=filename.replace('_','/')            
            message(path+file)
            try:
                f=open(path+file,'r',encoding='utf-8')
                doc=BeautifulSoup(f,"lxml")
                f.close()
                tables.extend(parse_table_mdpi(doc))
                ifsuccess=1
            except:
                ifsuccess=0
            if not ifsuccess:
                message("PRASE ERROR")
            doc={}
            doc['path']=path
            doc['file']=file
            doc['doi']=doi
            doc['content']=parser(path,file,'mdpi',doi).sections
            doc['tables']=tables
            doc_dict[doi]=doc
    log.close()
    return doc_dict


global log,iflog
iflog=0
#path='D:/glass/html_wiley/'
#path='D:/glass-total-results/html_mdpi/'
path='D:/glass-total-results/html_wiley/'
logpath=path+'parse.log'
#doc_dict=process_mdpi()
doc_dict=process_wiley()
# process_mdpi()
f=open(path+'wiley_data_strcut_20241010.json','w',encoding='utf-8')
#f=open(path+'mdpi_data_strcut_20241010.json','w',encoding='utf-8')
json.dump(doc_dict,f)
f.close()


# # test a single file
# f=open('E:/Data/Literature Data/AM fatigue html/mdpi/10.3390_app6020033.html','r',encoding='utf-8')
# doc=BeautifulSoup(f,"lxml")
# f.close()
# tables=parse_table_mdpi(doc)

# print_table(tables[0])
