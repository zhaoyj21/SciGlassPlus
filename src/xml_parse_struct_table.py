# -*- coding: utf-8 -*-
from xml.dom.minidom import parse
import json
import os
#from openpyxl import load_workbook,Workbook

def blank_line(text):
    ifblank=True
    for i in range(0,len(text)):
        if not ((text[i]==' ')or(text[i]=='\n')):
            ifblank=False
            return ifblank
    return ifblank

def parse_para(para):
    text=""
    if para.hasChildNodes():
        for c in para.childNodes:
            text1=parse_para(c)
            ifblank=blank_line(text1)
            if (len(text1)>0)and( not ifblank ):
                text=text+text1
    else:
        if hasattr(para,"data"):
            text=para.data
            #print(text)
    return text

def get_item(obj,tag):
    text=""
    lst=obj.getElementsByTagName(tag)
    if len(lst)==1:
        text=lst[0].childNodes[0].data
    return text


def parse_section(sec):
    seclst={'sec_title':'','content':[]}
    title=""
    textlst=[]
    idlst=[]
    sec_id=sec.getAttribute("id")
    ifmatch=False
    for cn in sec.childNodes:
        if hasattr(cn,"tagName"):
            tag=cn.tagName
            if tag=="ce:label":
                title=title+parse_para(cn)
            elif tag=="ce:section-title":
                title2=parse_para(cn)
                title=title+" "+title2
            elif tag=="ce:section":
                seclst1=parse_section(cn)
                seclst['content'].append(seclst1)
            elif tag=="ce:para":
                para_id=cn.getAttribute("id")
                para_text=parse_para(cn)
                seclst['content'].append(para_text)


    seclst['sec_title']=title         
    return seclst

def parse_doc(doc):
    
    sections=[]
    artflag=1       # Flag: if article exist, check by terms in VAR level
    txttype=""
    levels=["xocs:doc","xocs:serial-item","article","body","ce:sections"]
    tmp0=doc.getElementsByTagName("originalText")
    if not (len(tmp0)==1):
        print("originalText NOT EXIST")
        return 
    for level in levels:
        tmp1=tmp0[0].getElementsByTagName(level)
        if not (len(tmp1)==1):
            print(level+" NOT EXIST")
            artflag=0
            break
        tmp0=tmp1
    if artflag:
        cn=tmp0[0].childNodes
        for i in range(0,len(cn)):
            # parse section
            if hasattr(cn[i],'tagName')and(cn[i].tagName=='ce:section'):
                sec=parse_section(cn[i])
                sections.append(sec)
        txttype="article"
    else:
        tmp0=doc.getElementsByTagName("originalText")
        if not (len(tmp0)==1):
            return
        tmp1=tmp0[0].getElementsByTagName("xocs:doc")
        if not (len(tmp1)==1):
            return
        tmp2=tmp1[0].getElementsByTagName("xocs:rawtext")
        if not (len(tmp2)==1):
            return
        if hasattr(tmp2[0].childNodes[0],"data"):
            sections.append(tmp2[0].childNodes[0].data)
                    
    return sections

from xml.dom.minidom import parse
import json
import os
from openpyxl import load_workbook,Workbook
from prettytable import PrettyTable


def write_excel(path,doc):
    doi=doc['doi']
    name=doi.replace("/","_")
    tables=doc['table']
    
    wb=Workbook()
    for i in range(0,len(tables)):
        tab=tables[i]
        title=tab['title']
        head=tab['head']
        body=tab['body']
        legend=tab['legend']
        if i==0:
            ws=wb['Sheet']
            ws.title='1'
        else:
            ws=wb.create_sheet(str(i+1))
        ws.cell(1,1).value=title
        ct=1
        for j in range(0,len(head)):
            ct=ct+1
            for k in range(0,len(head[j])):
                ws.cell(ct,k+1).value=head[j][k]
        for j in range(0,len(body)):
            ct=ct+1
            for k in range(0,len(body[j])):
                ws.cell(ct,k+1).value=body[j][k]        
        if len(legend)>0:
            for j in range(0,len(legend)):
                ct=ct+1
                if isinstance(legend[j],str):
                    ws.cell(ct,1).value=legend[j]
                elif isinstance(legend[j],list):
                    for k in range(0,len(legend[j])):
                        ws.cell(ct,k+1).value=legend[j][k] 
    if len(tables)>0:
        wb.save(path+name+'.xlsx')
    wb.close()    


def blank_line(text):
    ifblank=True
    for i in range(0,len(text)):
        if not ((text[i]==' ')or(text[i]=='\n')):
            ifblank=False
            return ifblank
    return ifblank

def parse_para(para):
    text=""
    if para.hasChildNodes():
        for c in para.childNodes:
            text1=parse_para(c)
            ifblank=blank_line(text1)
            if (len(text1)>0)and( not ifblank ):
                text=text+text1
    else:
        if hasattr(para,"data"):
            text=para.data
            #print(text)
    return text

def get_item(obj,tag):
    text=""
    lst=obj.getElementsByTagName(tag)
    if len(lst)==1:
        text=lst[0].childNodes[0].data
    return text


def parse_rows(rows,ncol):
    data=[]
    nr=0
    cache=[]
    ntimes=[]
    pos=[]
    for r in rows:
        data.append([])
        if ncol>0:
            for i in range(0,ncol):
                data[-1].append('')
        entries=r.getElementsByTagName('entry')
        ct=0
        for i in range(0,len(cache)):
            if ntimes[i]>0:
                for p in pos[i]:
                    while len(data[-1])<p+1:
                        data[-1].append('')
                    data[-1][p]=cache[i]
                ntimes[i]=ntimes[i]-1
        for en in entries:
            namest=en.getAttribute('namest')
            nameend=en.getAttribute('nameend')
            morerows=en.getAttribute('morerows')
            keys=list(en._get_attributes().keys())
            if not morerows=='':
                cache.append(parse_para(en))
                ntimes.append(int(morerows))   
                pos.append([])                    
            if not (namest=='' and nameend==''):
                sid=int(namest[3:])
                eid=int(nameend[3:])
                tmp=parse_para(en)
                for i in range(sid-1,eid):
                    if not morerows=='':
                        pos[-1].append(i)
                    while len(data[-1])<i+1:
                        data[-1].append('')
                    data[-1][i]=tmp
                ct=eid
            else:
                if not morerows=='':
                    pos[-1].append(ct)
                while (len(data[-1])>=ct+1)and(not data[-1][ct]==''):
                    ct=ct+1
                while len(data[-1])<ct+1:
                    data[-1].append('')                
                data[-1][ct]=parse_para(en)
                ct=ct+1
    
    return data

def parse_tgroup(tg):
    ncol=0
    tmp=tg.getAttribute('ncol')
    if len(tmp)>0:
        ncol=int(tmp)
    tmp=tg.getAttribute('cols')
    if len(tmp)>0:
        ncol=int(tmp) 
    
    # process thead
    head=[]
    theads=tg.getElementsByTagName('thead')
    if len(theads)==1:
        th=theads[0]
        rows=th.getElementsByTagName('row')
        head=parse_rows(rows,ncol)
    
    # process tbody
    body=[]
    theads=tg.getElementsByTagName('tbody')
    if len(theads)==1:
        th=theads[0]
        rows=th.getElementsByTagName('row')
        body=parse_rows(rows,ncol)
    
    return head,body

def parse_table(sec):
    table={'title':'','head':[],'body':[],'legend':[]}
    title=''
    label=''
    caption=''
    legend=''
    head=[]
    body=[]
    sec_id=sec.getAttribute("id")
    ifmatch=False
    for cn in sec.childNodes:
        if hasattr(cn,"tagName"):
            tag=cn.tagName
            if tag=="ce:label":
                label=title+parse_para(cn)
            elif tag=="ce:caption":
                caption=parse_para(cn)
                
            elif tag=="tgroup":
                head,body=parse_tgroup(cn)
            elif tag=="ce:legend":
                legend=parse_para(cn)

    title=label+' '+caption
    table['title']=title
    table['head']=head
    table['body']=body         
    return table

def parse_doc_table(doc):
    global log
    tables=[]
    artflag=1       # Flag: if article exist, check by terms in VAR level
    txttype=""
    levels=["xocs:doc","xocs:serial-item","article","ce:floats"]
    tmp0=doc.getElementsByTagName("originalText")
    if not (len(tmp0)==1):
        print("originalText NOT EXIST")
        log.write('originalText NOT EXIST\n')
        return tables
    for level in levels:
        tmp1=tmp0[0].getElementsByTagName(level)
        if not (len(tmp1)==1):
            print(level+" NOT EXIST")
            log.write(level+" NOT EXIST\n")
            artflag=0
            break
        tmp0=tmp1
    if artflag:
        cn=tmp0[0].childNodes
        for i in range(0,len(cn)):
            # parse table
            if hasattr(cn[i],'tagName')and(cn[i].tagName=='ce:table'):
                tab=parse_table(cn[i])
                tables.append(tab)
        txttype="article"
                    
    return tables

# def get_doi_dict():
#     doi_dict={}
#     wospath="D:/Tsinghua/Project/Fatigue Data Framework/search data/wos/20220917/"
#     fileName="AM_fatigue_adddoi.xlsx"
#     sheetName="savedrecs"
#     wb=load_workbook(filename=wospath+fileName)
#     ws=wb[sheetName]
#     nrow=ws.max_row
#     ncol=ws.max_column
#     for i in range(1,ncol+1):
#         if ws.cell(1,i).value=='DOI':
#             col=i
#             break
#     for i in range(2,nrow+1):
#         doi=ws.cell(i,col).value.strip('\n')
#         doi_trans=doi.replace('/','_').replace(':','_')
#         doi_dict[doi_trans]=doi

#     return doi_dict


#path="E:/Data/Literature Data/glass/elsevier/xml/"
#logpath='E:/Data/Literature Data/glass/elsevier/tmp.log'
#path="D:/glass/xml_els/"
#logpath='D:/glass/xml_els_parse.log'
path="D:/glass-total-results/xml_els/"
logpath='D:/glass-total-results/xml_els_parse.log'
# doi_dict=get_doi_dict()
filelist=os.listdir(path)
doc_dict={}

log=open(logpath,'w')
for file in filelist:
    leng=len(file)
    if leng>4 and file[leng-4:leng]=='.xml':    
        filename=file[0:leng-4]
        #doi=doi_dict[filename]
        doi=filename.replace('_','/')
        print(file)
        log.write(file)
        try:   
            dom=parse(path+filename+'.xml')
            data=dom.documentElement
            sections=parse_doc(data)
            tables=parse_doc_table(data)
            log.write('\n')
        except:
            print('PARSE ERROR')
            log.write(' PRASE ERROR\n')
    doc={}
    doc['path']=path
    doc['file']=filename+'.xml'
    doc['doi']=doi
    doc['content']=sections
    doc['table']=tables
    doc_dict[doi]=doc

log.close()
    
f=open(path+'data_strcut_els_20241010.json','w',encoding='utf-8')
json.dump(doc_dict,f)
f.close()