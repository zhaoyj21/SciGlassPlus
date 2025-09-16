# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup
import os
import json
from openpyxl import load_workbook,Workbook
import requests
import time
import random


def get_doi_dict():
    doi_dict={}
    wospath="D:/Tsinghua/Project/Fatigue Data Framework/search data/wos/20220917/"
    fileName="AM_fatigue_adddoi.xlsx"
    sheetName="savedrecs"
    wb=load_workbook(filename=wospath+fileName)
    ws=wb[sheetName]
    nrow=ws.max_row
    ncol=ws.max_column
    for i in range(1,ncol+1):
        if ws.cell(1,i).value=='DOI':
            col=i
            break
    for i in range(2,nrow+1):
        doi=ws.cell(i,col).value.strip('\n')
        doi_trans=doi.replace('/','_').replace(':','_')
        doi_dict[doi_trans]=doi

    return doi_dict

# path='E:/Data/Literature Data/AM fatigue html/Springer/'
#path='D:/glass/html_springer/'
#path='D:/glass1419/html_springer/'
path='D:/glass-total-results/html_springer/'
# logpath='E:/Data/Literature Data/AM fatigue html/springer_table_download.log'
#logpath='D:/glass/html_springer/download.log'
#logpath='D:/glass1419/html_springer/download.log'
logpath='D:/glass-total-results/html_springer/download.log'
doi_dict={}
# doi_dict=get_doi_dict()

filelist=os.listdir(path)
filelist2=os.listdir(path+'table')
doc_dict={}
iftxt=0
log=open(logpath,'w')
log.close()
headers = {
    'Accept': 'text/html',
    'User-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36 Edg/97.0.1072.55'
}

for file in filelist[1159:]:
    log=open(logpath,'a+')
    leng=len(file)
    if leng>5 and file[leng-5:leng]=='.html':    
        filename=file[0:leng-5]
        if filename in filelist2:
            continue
        print(file)
        log.write(file+'\n')
        
        if not os.path.exists(path+'table/'+filename):
            os.mkdir(path+'table/'+filename)
        
        f=open(path+file,'r',encoding='utf-8')
        doc=BeautifulSoup(f,'lxml')
        f.close()
        try:
            tmp=doc.find('meta',{'property':{'og:url'}})
            if tmp.has_attr('content'):
                url0=tmp['content']
            else:
                print('NO URL')
                log.write('NO URL\n')
                continue
            tables=doc.find_all('div',class_='c-article-table')
            ct=0
            for tab in tables:
                ct=ct+1
                if tab.has_attr('id'):
                    tid=tab.attrs['id']
                else:
                    tid='TABLE'+str(ct)
                print(tid)
                log.write(tid+'\n')
                but=tab.find('a',class_='c-article__pill-button')
                if but.has_attr('href'):
                    url2=but.attrs['href']
                    fout=open(path+'table/'+filename+'/'+tid+'.html','wb')
                    r = requests.get('https://link.springer.com/'+url2,stream=True,headers=headers)
                    if r.status_code == 200:
                        for chunk in r.iter_content(2048):
                            fout.write(chunk)  
                    else:
                        print('ERROR') 
                        log.write('ERROR\n')
                    lag=random.uniform(3,10)
                    time.sleep(lag) 
        except:
            print('ERROR') 
            log.write('ERROR\n')            
    log.close()

# f=open(path+'data.json','w',encoding='utf-8')
# json.dump(doc_dict,f)
# f.close()